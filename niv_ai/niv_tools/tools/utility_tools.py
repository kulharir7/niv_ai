import frappe
import json
import requests
from frappe.utils import get_files_path


def get_doctype_info(doctype):
    """Get metadata about a DocType including fields, permissions, and workflow"""
    meta = frappe.get_meta(doctype)

    fields = []
    for f in meta.fields:
        field_info = {
            "fieldname": f.fieldname,
            "label": f.label,
            "fieldtype": f.fieldtype,
            "reqd": f.reqd,
            "options": f.options,
            "default": f.default,
        }
        if f.fieldtype == "Link":
            field_info["linked_doctype"] = f.options
        if f.fieldtype == "Select" and f.options:
            field_info["select_options"] = f.options.split("\n")
        fields.append(field_info)

    permissions = []
    for p in meta.permissions:
        permissions.append({
            "role": p.role,
            "read": p.read,
            "write": p.write,
            "create": p.create,
            "delete": p.delete,
            "submit": p.submit,
        })

    result = {
        "doctype": doctype,
        "module": meta.module,
        "is_submittable": meta.is_submittable,
        "is_tree": meta.is_tree,
        "is_single": meta.issingle,
        "is_child_table": meta.istable,
        "autoname": meta.autoname,
        "name_case": meta.name_case,
        "title_field": meta.title_field,
        "image_field": meta.image_field,
        "search_fields": meta.search_fields,
        "sort_field": meta.sort_field,
        "sort_order": meta.sort_order,
        "fields": fields,
        "permissions": permissions,
        "field_count": len(fields),
    }

    # Check workflow
    workflow = frappe.get_all("Workflow", filters={"document_type": doctype, "is_active": 1}, limit=1)
    if workflow:
        wf = frappe.get_doc("Workflow", workflow[0].name)
        result["workflow"] = {
            "name": wf.name,
            "states": [{"state": s.state, "doc_status": s.doc_status, "allow_edit": s.allow_edit} for s in wf.states],
            "transitions": [{"state": t.state, "action": t.action, "next_state": t.next_state, "allowed": t.allowed} for t in wf.transitions],
        }

    return result


def fetch_url(url, method="GET", headers=None, body=None, timeout=10):
    """Fetch content from a URL"""
    if not url.startswith(("http://", "https://")):
        return {"error": "URL must start with http:// or https://"}

    try:
        kwargs = {
            "url": url,
            "method": method.upper(),
            "timeout": min(timeout, 30),
            "headers": headers or {},
        }
        if body and method.upper() in ("POST", "PUT", "PATCH"):
            kwargs["data"] = json.dumps(body) if isinstance(body, dict) else body
            kwargs["headers"]["Content-Type"] = kwargs["headers"].get("Content-Type", "application/json")

        resp = requests.request(**kwargs)

        content_type = resp.headers.get("Content-Type", "")
        if "json" in content_type:
            response_body = resp.json()
        else:
            response_body = resp.text[:10000]  # Limit response size

        return {
            "status_code": resp.status_code,
            "content_type": content_type,
            "body": response_body,
        }
    except requests.Timeout:
        return {"error": "Request timed out"}
    except requests.RequestException as e:
        return {"error": str(e)}


def extract_file_content(file_url):
    """Extract text content from an uploaded file"""
    if not file_url:
        return {"error": "No file URL provided"}

    # Get the file doc
    files = frappe.get_all("File", filters={"file_url": file_url}, fields=["name", "file_name", "file_url", "is_private"], limit=1)
    if not files:
        return {"error": f"File not found: {file_url}"}

    file_doc = files[0]
    file_path = frappe.get_site_path(
        "private" if file_doc.is_private else "public",
        "files",
        file_doc.file_name
    )

    import os
    if not os.path.exists(file_path):
        return {"error": f"File not found on disk: {file_path}"}

    ext = os.path.splitext(file_doc.file_name)[1].lower()

    try:
        if ext in (".txt", ".md", ".csv", ".log", ".py", ".js", ".html", ".css", ".json", ".xml", ".yml", ".yaml"):
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read(50000)  # Limit to 50KB
            return {
                "file_name": file_doc.file_name,
                "content_type": "text",
                "content": content,
                "truncated": len(content) >= 50000,
            }

        elif ext == ".pdf":
            try:
                import PyPDF2
                with open(file_path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    pages = []
                    for i, page in enumerate(reader.pages[:50]):
                        text = page.extract_text()
                        if text:
                            pages.append(f"--- Page {i+1} ---\n{text}")
                return {
                    "file_name": file_doc.file_name,
                    "content_type": "pdf",
                    "pages": len(reader.pages),
                    "content": "\n\n".join(pages)[:50000],
                }
            except ImportError:
                return {"error": "PyPDF2 not installed. Cannot extract PDF content."}

        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = {}
                for sheet_name in wb.sheetnames[:10]:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(max_row=200, values_only=True):
                        rows.append([str(cell) if cell is not None else "" for cell in row])
                    sheets[sheet_name] = rows
                wb.close()
                return {
                    "file_name": file_doc.file_name,
                    "content_type": "spreadsheet",
                    "sheets": sheets,
                }
            except ImportError:
                return {"error": "openpyxl not installed. Cannot extract Excel content."}

        elif ext in (".docx",):
            try:
                import docx
                doc = docx.Document(file_path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                return {
                    "file_name": file_doc.file_name,
                    "content_type": "docx",
                    "content": "\n\n".join(paragraphs)[:50000],
                }
            except ImportError:
                return {"error": "python-docx not installed. Cannot extract Word content."}

        else:
            return {
                "file_name": file_doc.file_name,
                "content_type": "unsupported",
                "message": f"Cannot extract content from {ext} files",
            }

    except Exception as e:
        return {"error": f"Failed to extract content: {str(e)}"}


def create_dashboard(dashboard_name, charts=None):
    """Create a Frappe dashboard"""
    if frappe.db.exists("Dashboard", dashboard_name):
        return {"error": f"Dashboard '{dashboard_name}' already exists"}

    doc = frappe.get_doc({
        "doctype": "Dashboard",
        "name": dashboard_name,
        "dashboard_name": dashboard_name,
        "owner": frappe.session.user,
    })

    if charts:
        for chart in charts:
            doc.append("charts", {
                "chart": chart.get("chart"),
                "width": chart.get("width", "Full"),
            })

    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {
        "success": True,
        "name": doc.name,
        "message": f"Dashboard '{dashboard_name}' created successfully",
    }


def create_dashboard_chart(chart_name, chart_type="Count", doctype=None, based_on=None,
                           timespan="Last Year", time_interval="Monthly",
                           filters_json=None, group_by=None, aggregate_function="Count",
                           value_based_on=None):
    """Create a Dashboard Chart"""
    if frappe.db.exists("Dashboard Chart", chart_name):
        return {"error": f"Chart '{chart_name}' already exists"}

    chart_doc = {
        "doctype": "Dashboard Chart",
        "chart_name": chart_name,
        "chart_type": chart_type,
        "owner": frappe.session.user,
    }

    if chart_type in ("Count", "Sum", "Average"):
        chart_doc.update({
            "document_type": doctype,
            "based_on": based_on or "creation",
            "timespan": timespan,
            "time_interval": time_interval,
            "filters_json": json.dumps(filters_json or []),
            "group_by_type": "Count" if not group_by else "Group By",
            "group_by_based_on": group_by,
            "aggregate_function_based_on": value_based_on,
        })

    if chart_type == "Group By":
        chart_doc.update({
            "document_type": doctype,
            "group_by_type": "Count",
            "group_by_based_on": group_by or based_on,
            "aggregate_function_based_on": value_based_on,
        })

    doc = frappe.get_doc(chart_doc)
    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {
        "success": True,
        "name": doc.name,
        "message": f"Chart '{chart_name}' created successfully",
    }


def list_user_dashboards():
    """List all dashboards accessible to the current user"""
    dashboards = frappe.get_all(
        "Dashboard",
        fields=["name", "dashboard_name", "owner", "creation", "modified"],
        order_by="modified DESC",
        limit=50,
    )

    result = []
    for d in dashboards:
        charts = frappe.get_all(
            "Dashboard Chart Link",
            filters={"parent": d.name},
            fields=["chart", "width"],
        )
        d["charts"] = charts
        d["chart_count"] = len(charts)
        result.append(d)

    return {
        "count": len(result),
        "dashboards": result,
    }
