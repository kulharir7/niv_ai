"""
Export API — Generate Excel/CSV files from chat data.
Called from UI when user clicks download button on table/data responses.
"""
import csv
import io
import json
import os
import re
import uuid

import frappe
from frappe import _


@frappe.whitelist(allow_guest=False)
def export_data(data=None, format="excel", filename=None):
    """Export structured data to Excel, CSV, or PDF file.
    
    Args:
        data: JSON string — either:
            - A list of dicts: [{"col1": "val1", "col2": "val2"}, ...]
            - A markdown table string (auto-parsed)
        format: "excel", "csv", or "pdf"
        filename: Optional custom filename (without extension)
    
    Returns:
        {"file_url": "/files/export_xxx.xlsx", "filename": "export_xxx.xlsx"}
    """
    if not data:
        frappe.throw(_("No data provided for export"))

    # Parse data
    if isinstance(data, str):
        try:
            parsed = json.loads(data)
        except (json.JSONDecodeError, TypeError):
            # Try parsing as markdown table
            parsed = _parse_markdown_table(data)
            if not parsed:
                frappe.throw(_("Could not parse data. Send JSON array or markdown table."))
    else:
        parsed = data

    if not parsed or not isinstance(parsed, list) or len(parsed) == 0:
        frappe.throw(_("No rows found in data"))

    # Ensure all items are dicts
    if isinstance(parsed[0], dict):
        rows = parsed
    elif isinstance(parsed[0], (list, tuple)):
        # List of lists — first row is headers
        headers = [str(h) for h in parsed[0]]
        rows = []
        for row in parsed[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i}"
                row_dict[key] = val
            rows.append(row_dict)
    else:
        frappe.throw(_("Data must be a list of objects or list of lists"))

    if not rows:
        frappe.throw(_("No data rows to export"))

    # Get column headers from first row
    headers = list(rows[0].keys())

    # Generate filename
    ext = "xlsx" if format == "excel" else ("pdf" if format == "pdf" else "csv")
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', filename or "export")
    out_filename = f"{safe_name}_{uuid.uuid4().hex[:8]}.{ext}"

    # Output path
    output_dir = frappe.get_site_path("public", "files")
    output_path = os.path.join(output_dir, out_filename)

    if format == "excel":
        _write_excel(output_path, headers, rows)
    elif format == "pdf":
        _write_pdf(output_path, headers, rows, filename or "Export")
    else:
        _write_csv(output_path, headers, rows)

    file_url = f"/files/{out_filename}"

    return {
        "file_url": file_url,
        "filename": out_filename,
    }


def _write_excel(path, headers, rows):
    """Write data to Excel file using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")

            # Try to convert numeric strings to numbers for proper Excel formatting
            if isinstance(value, str):
                value = value.strip()
                try:
                    if "." in value:
                        value = float(value)
                    elif value.isdigit():
                        value = int(value)
                except (ValueError, AttributeError):
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit column widths (approximate)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_data in rows[:100]:  # Sample first 100 rows
            val = str(row_data.get(header, ""))
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    wb.save(path)


def _write_pdf(path, headers, rows, title="Export"):
    """Write data to PDF file using wkhtmltopdf (via pdfkit or subprocess)."""
    from datetime import datetime

    # Build HTML table
    now = datetime.now().strftime("%d %b %Y, %I:%M %p")
    user = frappe.session.user
    user_name = frappe.db.get_value("User", user, "full_name") or user

    header_cells = "".join(f"<th>{_escape_html(str(h))}</th>" for h in headers)

    body_rows = ""
    for i, row in enumerate(rows):
        cls = ' class="alt"' if i % 2 == 0 else ""
        cells = "".join(
            f"<td>{_escape_html(str(row.get(h, '')))}</td>" for h in headers
        )
        body_rows += f"<tr{cls}>{cells}</tr>\n"

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
    body {{ font-family: Arial, sans-serif; font-size: 12px; color: #333; margin: 20px; }}
    h2 {{ color: #1a237e; margin-bottom: 4px; }}
    .meta {{ color: #666; font-size: 10px; margin-bottom: 16px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
    th {{ background: #4472C4; color: #fff; font-weight: 600; text-transform: uppercase;
         font-size: 10px; letter-spacing: 0.03em; padding: 8px 10px; text-align: left;
         border: 1px solid #3460a8; }}
    td {{ padding: 7px 10px; border: 1px solid #ddd; font-size: 11px; }}
    tr.alt td {{ background: #f2f6fc; }}
    tr:hover td {{ background: #e8eef8; }}
    .footer {{ margin-top: 20px; font-size: 9px; color: #999; text-align: right; }}
</style>
</head>
<body>
    <h2>{_escape_html(title)}</h2>
    <div class="meta">Generated by Niv AI &middot; {now} &middot; {_escape_html(user_name)}</div>
    <table>
        <thead><tr>{header_cells}</tr></thead>
        <tbody>{body_rows}</tbody>
    </table>
    <div class="footer">{len(rows)} rows &middot; Niv AI Export</div>
</body>
</html>"""

    # Use pdfkit (wkhtmltopdf wrapper) or subprocess
    try:
        import pdfkit
        pdfkit.from_string(html, path, options={
            "page-size": "A4",
            "orientation": "Landscape" if len(headers) > 5 else "Portrait",
            "margin-top": "10mm",
            "margin-bottom": "10mm",
            "margin-left": "10mm",
            "margin-right": "10mm",
            "encoding": "UTF-8",
            "quiet": "",
        })
    except ImportError:
        # Fallback: direct wkhtmltopdf via subprocess
        import subprocess
        import tempfile

        tmp_html = tempfile.NamedTemporaryFile(suffix=".html", delete=False, mode="w", encoding="utf-8")
        tmp_html.write(html)
        tmp_html.close()

        orientation = "Landscape" if len(headers) > 5 else "Portrait"
        try:
            subprocess.run(
                ["wkhtmltopdf", "--quiet", "--page-size", "A4",
                 "--orientation", orientation,
                 "--margin-top", "10mm", "--margin-bottom", "10mm",
                 "--margin-left", "10mm", "--margin-right", "10mm",
                 "--encoding", "UTF-8",
                 tmp_html.name, path],
                check=True, timeout=30,
            )
        finally:
            try:
                os.unlink(tmp_html.name)
            except Exception:
                pass


def _escape_html(text):
    """Escape HTML special characters."""
    return (text
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;"))


def _write_csv(path, headers, rows):
    """Write data to CSV file."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _parse_markdown_table(text):
    """Parse a markdown table string into list of dicts.
    
    Example input:
        | Name | Amount |
        |------|--------|
        | ABC  | 1000   |
        | XYZ  | 2000   |
    
    Returns: [{"Name": "ABC", "Amount": "1000"}, {"Name": "XYZ", "Amount": "2000"}]
    """
    if not text or "|" not in text:
        return None

    lines = [line.strip() for line in text.strip().split("\n") if line.strip()]

    # Filter only table lines (start and end with |)
    table_lines = [l for l in lines if l.startswith("|") and l.endswith("|")]
    if len(table_lines) < 2:
        return None

    # Parse header
    header_line = table_lines[0]
    headers = [h.strip() for h in header_line.strip("|").split("|") if h.strip()]
    if not headers:
        return None

    # Skip separator row (|---|---|)
    data_lines = []
    for line in table_lines[1:]:
        cells = [c.strip() for c in line.strip("|").split("|")]
        # Skip if all cells are just dashes (separator row)
        if all(re.match(r'^[-:]+$', c) for c in cells if c):
            continue
        data_lines.append(cells)

    if not data_lines:
        return None

    # Build dicts
    result = []
    for cells in data_lines:
        row = {}
        for i, header in enumerate(headers):
            row[header] = cells[i].strip() if i < len(cells) else ""
        result.append(row)

    return result
